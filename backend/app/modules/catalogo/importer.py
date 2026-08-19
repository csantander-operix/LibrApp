"""Importador masivo de catálogo desde Excel/CSV (RF-05 / CU-04 / HU-02).

Diseñado para tolerar los archivos reales de la librería, que son planillas
operativas (no exports limpios):
  - Filas de título/metadata antes del encabezado.
  - Delimitador ';' (Excel en español) o ','.
  - Precios en formato argentino: "$ 25.000,00".
  - Columnas faltantes (ej: sin autor, sin editorial, sin ubicación).

Reglas aplicadas:
  - RN-02: si falta autor/editorial se completa con "Sin especificar" (no se rechaza la fila).
  - RN-07/RN-09: sin estante o estante inexistente → el libro queda 'Sin ubicar'.
  - RN-01 / P-04: si el ISBN ya existe, se ACTUALIZA el registro (upsert); si no, se crea.
  - RN-10: si viene una colección que no existe, se crea.
"""
import csv
import io
import unicodedata
from decimal import Decimal, InvalidOperation

from sqlalchemy.orm import Session

from app.modules.catalogo.models import Coleccion, Estante, Libro
from app.modules.catalogo.schemas import ImportResultado, ImportFilaError

# Sinónimos aceptados por campo destino (comparados sobre el header normalizado).
_ALIAS = {
    "titulo": {"titulo", "title", "nombre"},
    "autor": {"autor", "autores", "author"},
    "editorial": {"editorial", "edit"},
    "isbn": {"isbn", "isbn13", "cod", "codigo"},
    "precio": {"precio", "price", "importe"},
    "coleccion": {"coleccion", "categoria", "collection"},
    "estante": {"estante", "ubicacion", "estantecodigo", "estante_codigo", "sector", "seccion"},
}


def _norm(s: str) -> str:
    """Minúsculas, sin acentos, sin espacios/puntos, para comparar encabezados."""
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return "".join(ch for ch in s.lower() if ch.isalnum())


def _detectar_campo(header: str) -> str | None:
    n = _norm(header)
    if not n:
        return None
    for campo, alias in _ALIAS.items():
        if n in alias or any(n.startswith(a) for a in alias):
            return campo
    return None


def parse_precio(raw) -> Decimal | None:
    """Convierte "$ 25.000,00" / "25000" / "25.5" a Decimal. None si no se puede."""
    if raw is None:
        return None
    if isinstance(raw, (int, float, Decimal)):
        try:
            return Decimal(str(raw))
        except InvalidOperation:
            return None
    s = str(raw).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(" ", "").replace("\xa0", "")
    tiene_coma = "," in s
    tiene_punto = "." in s
    if tiene_coma and tiene_punto:
        # Formato AR: punto = miles, coma = decimales.
        s = s.replace(".", "").replace(",", ".")
    elif tiene_coma:
        s = s.replace(",", ".")
    elif tiene_punto:
        # "25.000" (miles) vs "25.5" (decimal): si lo que sigue al último punto
        # son 3 dígitos, lo tratamos como separador de miles.
        entero, _, dec = s.rpartition(".")
        if len(dec) == 3 and entero.isdigit():
            s = entero + dec
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _leer_filas(contenido: bytes, filename: str) -> list[list[str]]:
    """Devuelve todas las filas como listas de strings, sea CSV o XLSX."""
    nombre = (filename or "").lower()
    if nombre.endswith(".xlsx") or nombre.endswith(".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
        filas: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            filas.append(["" if c is None else str(c) for c in row])
        wb.close()
        return filas
    # CSV: decodificar tolerante y detectar delimitador.
    texto = None
    for enc in ("utf-8-sig", "latin-1"):
        try:
            texto = contenido.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        texto = contenido.decode("utf-8", errors="replace")
    muestra = "\n".join(texto.splitlines()[:20])
    delim = ";" if muestra.count(";") >= muestra.count(",") else ","
    return [row for row in csv.reader(io.StringIO(texto), delimiter=delim)]


def _encontrar_header(filas: list[list[str]]) -> tuple[int, dict[str, int]]:
    """Busca la fila de encabezado (la que tiene una columna 'título') en las
    primeras filas. Devuelve (índice_fila, {campo: índice_columna})."""
    for i, fila in enumerate(filas[:15]):
        mapeo: dict[str, int] = {}
        for j, celda in enumerate(fila):
            campo = _detectar_campo(celda)
            if campo and campo not in mapeo:
                mapeo[campo] = j
        if "titulo" in mapeo:
            return i, mapeo
    raise ValueError(
        "No se encontró una columna de 'título' en el archivo. "
        "Revisá que tenga un encabezado con al menos la columna Título."
    )


def importar(db: Session, contenido: bytes, filename: str, dry_run: bool) -> ImportResultado:
    filas = _leer_filas(contenido, filename)
    if not filas:
        raise ValueError("El archivo está vacío")

    header_idx, mapeo = _encontrar_header(filas)
    columnas_detectadas = {
        campo: (filas[header_idx][idx].strip() or f"col{idx}") for campo, idx in mapeo.items()
    }

    # Índices de estantes y colecciones existentes (por nombre/código normalizado).
    estantes = {_norm(e.codigo): e for e in db.query(Estante).all()}
    colecciones = {_norm(c.nombre): c for c in db.query(Coleccion).all()}

    # Índices de libros existentes para el upsert. Por ISBN (RN-01) y, para los
    # libros sin ISBN (buena parte del inventario real), por (título+editorial):
    # así reimportar el mismo archivo actualiza en vez de duplicar el catálogo.
    por_isbn: dict[str, Libro] = {}
    por_titulo: dict[tuple[str, str], Libro] = {}
    for lb in db.query(Libro).all():
        if lb.isbn:
            por_isbn[lb.isbn] = lb
        else:
            por_titulo[(_norm(lb.titulo), _norm(lb.editorial))] = lb

    creados = actualizados = sin_ubicar = 0
    total_filas = 0
    errores: list[ImportFilaError] = []

    def celda(fila: list[str], campo: str) -> str | None:
        idx = mapeo.get(campo)
        if idx is None or idx >= len(fila):
            return None
        val = (fila[idx] or "").strip()
        return val or None

    for n, fila in enumerate(filas[header_idx + 1:], start=header_idx + 2):
        if not any((c or "").strip() for c in fila):
            continue  # fila totalmente vacía
        titulo = celda(fila, "titulo")
        if not titulo:
            continue  # sin título (probable fila de subtotal/metadata): se ignora
        total_filas += 1

        isbn = celda(fila, "isbn")
        autor = celda(fila, "autor") or "Sin especificar"
        editorial = celda(fila, "editorial") or "Sin especificar"
        precio = parse_precio(celda(fila, "precio"))

        # Colección: se resuelve o se crea (RN-10).
        coleccion = None
        nom_col = celda(fila, "coleccion")
        if nom_col:
            coleccion = colecciones.get(_norm(nom_col))
            if not coleccion and not dry_run:
                coleccion = Coleccion(nombre=nom_col)
                db.add(coleccion)
                db.flush()
                colecciones[_norm(nom_col)] = coleccion

        # Estante: match por código; si no existe → Sin ubicar (RN-09).
        estante = None
        cod_est = celda(fila, "estante")
        if cod_est:
            estante = estantes.get(_norm(cod_est))
        if estante is None:
            sin_ubicar += 1

        # Upsert: por ISBN (RN-01 / P-04) o, sin ISBN, por (título+editorial).
        titulo_key = (_norm(titulo), _norm(editorial))
        existente = por_isbn.get(isbn) if isbn else por_titulo.get(titulo_key)
        if existente is not None:
            actualizados += 1
            if not dry_run:
                existente.titulo = titulo
                existente.autor = autor
                existente.editorial = editorial
                if precio is not None:
                    existente.precio = precio
                if coleccion is not None:
                    existente.coleccion_id = coleccion.id
                if estante is not None:
                    existente.estante_id = estante.id
        else:
            creados += 1
            nuevo = Libro(
                isbn=isbn, titulo=titulo, autor=autor, editorial=editorial, precio=precio,
                coleccion_id=coleccion.id if coleccion else None,
                estante_id=estante.id if estante else None,
            )
            if not dry_run:
                db.add(nuevo)
            # Registrar en los índices para no duplicar dentro del mismo archivo.
            if isbn:
                por_isbn[isbn] = nuevo
            else:
                por_titulo[titulo_key] = nuevo

    if dry_run:
        db.rollback()
    else:
        db.commit()

    return ImportResultado(
        dry_run=dry_run,
        total_filas=total_filas,
        creados=creados,
        actualizados=actualizados,
        sin_ubicar=sin_ubicar,
        errores=errores,
        columnas_detectadas=columnas_detectadas,
    )
